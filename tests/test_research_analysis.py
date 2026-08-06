import importlib.util
import io
import unittest
from unittest.mock import patch

import research_analysis


def fixture_analysis(item_id="research_fixture"):
    segments = [
        {
            "id": "s1",
            "start": 0.0,
            "end": 2.0,
            "duration": 2.0,
            "speaker": "A",
            "speaker_name": "参加者A",
            "role": "participant",
            "text": "改善案に賛成です。",
            "characters": 9,
            "question_candidate": False,
            "annotation": {
                "codes": ["improvement"],
                "interaction_tags": ["agreement"],
                "important": False,
                "excluded": False,
            },
            "excluded": False,
        },
        {
            "id": "s2",
            "start": 2.5,
            "end": 6.5,
            "duration": 4.0,
            "speaker": "B",
            "speaker_name": "参加者B",
            "role": "participant",
            "text": "操作方法を改善したいです。",
            "characters": 12,
            "question_candidate": False,
            "annotation": {
                "codes": ["improvement"],
                "interaction_tags": [],
                "important": True,
                "excluded": False,
            },
            "excluded": False,
        },
        {
            "id": "s3",
            "start": 7.0,
            "end": 8.0,
            "duration": 1.0,
            "speaker": "A",
            "speaker_name": "参加者A",
            "role": "participant",
            "text": "質問ですか？",
            "characters": 6,
            "question_candidate": True,
            "annotation": {
                "codes": [],
                "interaction_tags": [],
                "important": False,
                "excluded": False,
            },
            "excluded": False,
        },
    ]
    return {
        "schema_version": 1,
        "algorithm_version": "fixture",
        "generated_at": "2026-07-31T00:00:00+00:00",
        "item": {
            "id": item_id,
            "source_name": '=unsafe".wav',
            "revision_count": 1,
            "analysis_revision": 2,
            "analysis_updated_at": "2026-07-31T00:00:00+00:00",
        },
        "config": {
            "stop_words": [],
            "morph_split_mode": "C",
            "cooccurrence_min_count": 1,
            "cooccurrence_top_terms": 60,
            "statistics_group_by": "speaker",
        },
        "classification": {"automatic": [], "configured": [], "manual": []},
        "cautions": [],
        "automatic": {},
        "manual": {
            "codebook": [{"id": "improvement", "label": "改善"}],
        },
        "segments": segments,
        "exports": {
            "json": f"/api/library/{item_id}/analysis/export.json",
        },
    }


class ResearchAnalysisTests(unittest.TestCase):
    def setUp(self):
        research_analysis._RESEARCH_CACHE.clear()

    def test_fallback_keeps_analysis_available_and_labels_limit(self):
        analysis = fixture_analysis("fallback_fixture")
        with (
            patch.object(research_analysis, "_load_ginza", return_value=(None, "not installed")),
            patch.object(research_analysis, "_load_sudachi", return_value=(None, "not installed")),
        ):
            result = research_analysis.enrich_research_analysis(analysis)

        research = result["research"]
        self.assertEqual(research["linguistics"]["engine"]["status"], "fallback")
        self.assertGreater(research["linguistics"]["coverage"]["token_count"], 0)
        self.assertTrue(research["linguistics"]["morpheme_preview"])
        self.assertNotIn("morphemes", research["linguistics"])
        self.assertEqual(research["segments"], [])
        self.assertIn("xlsx", result["exports"])
        self.assertIn("morphemes", result["exports"])
        self.assertTrue(
            any("研究データとして使用できません" in value
                for value in [research["linguistics"]["engine"]["message"]])
        )

    def test_complete_rows_contain_spreadsheet_and_spss_style_datasets(self):
        analysis = fixture_analysis("complete_fixture")
        with (
            patch.object(research_analysis, "_load_ginza", return_value=(None, "not installed")),
            patch.object(research_analysis, "_load_sudachi", return_value=(None, "not installed")),
        ):
            result = research_analysis.enrich_research_analysis(
                analysis,
                include_rows=True,
            )

        sources = research_analysis.research_csv_sources(result)
        self.assertEqual(len(sources["segments_all"]), 3)
        self.assertTrue(sources["morphemes"])
        self.assertTrue(sources["term_frequency"])
        speaker_terms = result["research"]["linguistics"]["speaker_term_frequency"]
        self.assertEqual({row["speaker"] for row in speaker_terms}, {"A", "B"})
        self.assertTrue(all(row["terms"] for row in speaker_terms))
        self.assertTrue(sources["descriptives"])
        self.assertTrue(sources["frequencies"])
        self.assertTrue(sources["crosstabs"])
        self.assertTrue(sources["statistical_tests"])
        self.assertTrue(sources["correlations"])
        self.assertTrue(sources["analysis_methods"])
        duration = next(
            row for row in sources["descriptives"]
            if row["scope"] == "overall" and row["variable"] == "duration_seconds"
        )
        self.assertEqual(duration["n"], 3)
        self.assertAlmostEqual(duration["mean"], 7 / 3)

    def test_manually_selected_term_builds_crosstab_and_chi_square(self):
        analysis = fixture_analysis("selected_term_fixture")
        analysis["config"]["crosstab_terms"] = ["改善"]
        with (
            patch.object(research_analysis, "_load_ginza", return_value=(None, "not installed")),
            patch.object(research_analysis, "_load_sudachi", return_value=(None, "not installed")),
        ):
            result = research_analysis.enrich_research_analysis(analysis, include_rows=True)

        statistics = result["research"]["statistics"]
        self.assertEqual(statistics["selected_terms"], ["改善"])
        table_rows = [
            row for row in statistics["crosstabs"]
            if row["column_variable"] == "selected_term:改善"
        ]
        self.assertEqual(len(table_rows), 4)
        self.assertEqual({row["row_value"] for row in table_rows}, {"参加者A", "参加者B"})
        self.assertEqual({row["column_value"] for row in table_rows}, {"あり", "なし"})
        test = next(
            row for row in statistics["tests"]
            if row["outcome_label"] == "単語「改善」"
        )
        self.assertEqual(test["test"], "Pearsonのカイ二乗検定")
        self.assertEqual(test["effect_name"], "cramers_v")
        self.assertEqual(test["n"], 3)
        self.assertEqual(test["status"], "computed")

    def test_excel_cells_are_formula_safe(self):
        self.assertEqual(research_analysis._excel_value("=1+1"), "'=1+1")
        self.assertEqual(research_analysis._excel_value(" @SUM(A1)"), "' @SUM(A1)")
        self.assertEqual(research_analysis._excel_value(-2.5), -2.5)

    @unittest.skipUnless(
        importlib.util.find_spec("openpyxl"),
        "openpyxl is installed by requirements.txt",
    )
    def test_workbook_has_standard_research_sheets(self):
        from openpyxl import load_workbook

        analysis = fixture_analysis("workbook_fixture")
        with (
            patch.object(research_analysis, "_load_ginza", return_value=(None, "not installed")),
            patch.object(research_analysis, "_load_sudachi", return_value=(None, "not installed")),
        ):
            analysis = research_analysis.enrich_research_analysis(
                analysis,
                include_rows=True,
            )
        sources = research_analysis.research_csv_sources(analysis)
        datasets = {
            name: (research_analysis.RESEARCH_CSV_FIELDS[name], rows)
            for name, rows in sources.items()
        }
        content = research_analysis.build_analysis_workbook(analysis, datasets)
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=False)
        self.assertIn("README", workbook.sheetnames)
        self.assertIn("発話データ", workbook.sheetnames)
        self.assertIn("形態素", workbook.sheetnames)
        self.assertIn("構文・係り受け", workbook.sheetnames)
        self.assertIn("統計検定", workbook.sheetnames)
        source_value = workbook["README"]["B4"].value
        self.assertTrue(str(source_value).startswith("'="))


if __name__ == "__main__":
    unittest.main()
