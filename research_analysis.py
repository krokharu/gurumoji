"""Research-oriented Japanese text and statistical analysis.

The module is deliberately imported by the Flask application without importing
GiNZA, SudachiPy, SciPy, or openpyxl at module import time.  This keeps the web
UI available while optional analysis dependencies are being diagnosed.
"""

from __future__ import annotations

import io
import json
import math
import re
import threading
from collections import Counter, OrderedDict, defaultdict
from importlib import metadata
from itertools import combinations
from statistics import mean, median, stdev
from typing import Any, Callable


RESEARCH_ALGORITHM_VERSION = "research-ja-2"
CONTENT_UPOS = {"NOUN", "PROPN", "VERB", "ADJ", "ADV"}
CONTENT_POS_JA = {"名詞", "動詞", "形容詞", "副詞"}
DEFAULT_STOP_WORDS = {
    "これ", "それ", "あれ", "ここ", "そこ", "ため", "よう", "もの", "こと",
    "ところ", "そう", "さん", "する", "ある", "いる", "なる", "できる",
    "思う", "言う", "いう", "私", "僕", "わたし", "はい", "ええ", "まあ",
}
FALLBACK_TOKEN_PATTERN = re.compile(
    r"[一-龯々〆ヵヶ]+|[ぁ-ゖー]+|[ァ-ヺー]+|"
    r"[A-Za-z]+(?:[-_'][A-Za-z0-9]+)*|[0-9]+(?:[.,][0-9]+)*|[^\s]",
    re.UNICODE,
)
SENTENCE_END_PATTERN = re.compile(r"[。！？!?]+$")
EXCEL_FORMULA_PREFIXES = ("=", "+", "-", "@")

SOURCE_REFERENCES = [
    {
        "id": "khcoder",
        "title": "KH Coder 3 Reference Manual",
        "authors": "樋口耕一 / Koichi Higuchi",
        "url": "https://khcoder.net/en/manual_en_v3.pdf",
        "note": "抽出語、TF・DF、品詞選択、共起ネットワーク等の計量テキスト分析を設計する際の参照資料。本実装はKH Coder互換ではありません。",
    },
    {
        "id": "sudachipy",
        "title": "SudachiPy official documentation",
        "authors": "Works Applications",
        "url": "https://worksapplications.github.io/sudachi.rs/python/",
        "note": "日本語の形態素分割、品詞、正規形、辞書形、読みの取得方法。",
    },
    {
        "id": "ginza",
        "title": "GiNZA: Japanese NLP Library based on Universal Dependencies",
        "authors": "Megagon Labs",
        "url": "https://megagonlabs.github.io/ginza/",
        "note": "SudachiPyを用いた形態素情報とUniversal Dependencies準拠の日本語係り受け解析。",
    },
    {
        "id": "scipy",
        "title": "SciPy statistical functions",
        "authors": "SciPy community",
        "url": "https://docs.scipy.org/doc/scipy/reference/stats.html",
        "note": "カイ二乗検定、分散分析、Kruskal–Wallis検定、相関係数の実装。",
    },
]


RESEARCH_CSV_FIELDS: dict[str, list[str]] = {
    "segments_all": [
        "segment_id", "start", "end", "duration_seconds", "speaker",
        "speaker_name", "role", "text", "characters", "token_count",
        "content_token_count", "unique_content_terms", "lexical_diversity",
        "characters_per_minute", "question_candidate", "code_ids",
        "code_labels", "interaction_tags", "important", "excluded",
    ],
    "morphemes": [
        "segment_id", "speaker", "speaker_name", "role", "sentence_id",
        "token_id", "sentence_token_id", "surface", "lemma", "normalized",
        "reading", "upos", "pos_detail", "inflection", "begin", "end",
        "is_content", "is_stop", "excluded",
    ],
    "dependencies": [
        "segment_id", "speaker", "speaker_name", "sentence_id", "token_id",
        "surface", "lemma", "upos", "dependency", "head_token_id",
        "head_surface", "named_entity", "excluded",
    ],
    "pos_frequency": ["upos", "pos_detail", "count", "percent"],
    "term_frequency": [
        "rank", "term", "lemma", "upos", "term_frequency",
        "document_frequency", "speaker_frequency", "term_percent",
        "document_percent", "tf_idf",
    ],
    "cooccurrence": [
        "rank", "term_a", "term_b", "cooccurrence_count",
        "document_frequency_a", "document_frequency_b", "jaccard", "dice",
    ],
    "descriptives": [
        "scope", "group_variable", "group", "variable", "label", "n",
        "missing", "mean", "standard_deviation", "median", "minimum",
        "q1", "q3", "maximum",
    ],
    "frequencies": ["variable", "label", "value", "count", "percent"],
    "crosstabs": [
        "table_id", "table_label", "row_variable", "row_value",
        "column_variable", "column_value", "count", "row_percent",
        "column_percent", "total_percent",
    ],
    "statistical_tests": [
        "family", "test", "outcome", "outcome_label", "group_variable",
        "n", "groups", "statistic", "df1", "df2", "p_value",
        "effect_name", "effect_size", "significant_0_05", "status",
        "assumption_note",
    ],
    "correlations": [
        "method", "variable_a", "label_a", "variable_b", "label_b", "n",
        "coefficient", "p_value", "significant_0_05", "status",
    ],
    "analysis_methods": [
        "category", "method", "engine", "engine_version", "status",
        "analysis_unit", "description", "source_title", "source_url",
    ],
}


_ENGINE_LOCK = threading.RLock()
_GINZA_MODEL: Any | None = None
_GINZA_LOAD_ATTEMPTED = False
_GINZA_ERROR = ""
_SUDACHI_DICTIONARY: Any | None = None
_SUDACHI_LOAD_ATTEMPTED = False
_SUDACHI_ERROR = ""
_RESEARCH_CACHE: "OrderedDict[tuple[Any, ...], dict[str, Any]]" = OrderedDict()
_RESEARCH_CACHE_SIZE = 4


def _package_version(name: str) -> str:
    try:
        return metadata.version(name)
    except metadata.PackageNotFoundError:
        return ""


def _safe_error(error: BaseException) -> str:
    text = re.sub(r"\s+", " ", str(error)).strip()
    return text[:300] or error.__class__.__name__


def _finite(value: Any, digits: int = 8) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(number):
        return None
    return round(number, digits)


def _percentile(values: list[float], proportion: float) -> float | None:
    if not values:
        return None
    ordered = sorted(values)
    if len(ordered) == 1:
        return ordered[0]
    position = (len(ordered) - 1) * proportion
    lower = math.floor(position)
    upper = math.ceil(position)
    if lower == upper:
        return ordered[lower]
    weight = position - lower
    return ordered[lower] * (1 - weight) + ordered[upper] * weight


def _load_ginza() -> tuple[Any | None, str]:
    global _GINZA_MODEL, _GINZA_LOAD_ATTEMPTED, _GINZA_ERROR
    with _ENGINE_LOCK:
        if _GINZA_LOAD_ATTEMPTED:
            return _GINZA_MODEL, _GINZA_ERROR
        _GINZA_LOAD_ATTEMPTED = True
        try:
            import spacy

            try:
                _GINZA_MODEL = spacy.load(
                    "ja_ginza",
                    config={
                        "components": {
                            "compound_splitter": {"split_mode": "C"},
                        },
                    },
                )
            except Exception:
                # spaCy 3.8 validates GiNZA 5.2's legacy null split_mode more
                # strictly.  The dependency parser does not require the
                # compound/bunsetsu post-processing components.
                _GINZA_MODEL = spacy.load(
                    "ja_ginza",
                    exclude=["compound_splitter", "bunsetu_recognizer"],
                )
        except Exception as error:  # optional dependency and model diagnostics
            _GINZA_ERROR = _safe_error(error)
        return _GINZA_MODEL, _GINZA_ERROR


def _load_sudachi() -> tuple[Any | None, str]:
    global _SUDACHI_DICTIONARY, _SUDACHI_LOAD_ATTEMPTED, _SUDACHI_ERROR
    with _ENGINE_LOCK:
        if _SUDACHI_LOAD_ATTEMPTED:
            return _SUDACHI_DICTIONARY, _SUDACHI_ERROR
        _SUDACHI_LOAD_ATTEMPTED = True
        try:
            from sudachipy import Dictionary

            _SUDACHI_DICTIONARY = Dictionary(dict="core")
        except Exception as error:  # optional dependency and dictionary diagnostics
            _SUDACHI_ERROR = _safe_error(error)
        return _SUDACHI_DICTIONARY, _SUDACHI_ERROR


def _clean_stop_words(config: dict[str, Any]) -> set[str]:
    values = config.get("stop_words") if isinstance(config.get("stop_words"), list) else []
    return DEFAULT_STOP_WORDS | {
        str(value).strip().casefold() for value in values if str(value).strip()
    }


def _upos_from_sudachi(pos: tuple[str, ...]) -> str:
    primary = pos[0] if pos else ""
    mapping = {
        "名詞": "NOUN",
        "代名詞": "PRON",
        "動詞": "VERB",
        "形容詞": "ADJ",
        "形状詞": "ADJ",
        "副詞": "ADV",
        "連体詞": "DET",
        "接続詞": "CCONJ",
        "感動詞": "INTJ",
        "助詞": "ADP",
        "助動詞": "AUX",
        "接頭辞": "X",
        "接尾辞": "X",
        "記号": "PUNCT",
        "補助記号": "PUNCT",
        "空白": "SPACE",
    }
    return mapping.get(primary, "X")


def _row_common(segment: dict[str, Any]) -> dict[str, Any]:
    return {
        "segment_id": str(segment.get("id") or ""),
        "speaker": str(segment.get("speaker") or "UNKNOWN"),
        "speaker_name": str(
            segment.get("speaker_name") or segment.get("speaker") or "話者未判定"
        ),
        "role": str(segment.get("role") or "participant"),
        "excluded": bool(segment.get("excluded")),
    }


def _analyze_with_ginza(
    segments: list[dict[str, Any]],
    nlp: Any,
    stop_words: set[str],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], int]:
    morphemes: list[dict[str, Any]] = []
    dependencies: list[dict[str, Any]] = []
    sentence_count = 0
    with _ENGINE_LOCK:
        for segment in segments:
            text = str(segment.get("text") or "")
            if not text:
                continue
            doc = nlp(text)
            common = _row_common(segment)
            token_ids = {
                token.i: index
                for index, token in enumerate(
                    (value for value in doc if not value.is_space), 1
                )
            }
            sentence_ids: dict[int, tuple[int, int]] = {}
            for sentence_id, sentence in enumerate(doc.sents, 1):
                sentence_count += 1
                position = 0
                for token in sentence:
                    if token.is_space:
                        continue
                    position += 1
                    sentence_ids[token.i] = (sentence_id, position)
            for token in doc:
                if token.is_space:
                    continue
                sentence_id, sentence_position = sentence_ids.get(token.i, (1, token.i + 1))
                normalized = str(token.norm_ or token.lemma_ or token.text)
                lemma = str(token.lemma_ or normalized or token.text)
                reading_values = token.morph.get("Reading")
                inflection_values = token.morph.get("Inflection")
                is_stop = bool(token.is_stop) or normalized.casefold() in stop_words
                is_content = (
                    token.pos_ in CONTENT_UPOS
                    and not token.is_punct
                    and not is_stop
                )
                row = {
                    **common,
                    "sentence_id": sentence_id,
                    "token_id": token_ids.get(token.i, token.i + 1),
                    "sentence_token_id": sentence_position,
                    "surface": str(token.text),
                    "lemma": lemma,
                    "normalized": normalized,
                    "reading": "|".join(str(value) for value in reading_values),
                    "upos": str(token.pos_ or "X"),
                    "pos_detail": str(token.tag_ or ""),
                    "inflection": "|".join(str(value) for value in inflection_values),
                    "begin": int(token.idx),
                    "end": int(token.idx + len(token.text)),
                    "is_content": is_content,
                    "is_stop": is_stop,
                }
                morphemes.append(row)
                head_is_root = token.dep_.upper() == "ROOT" or token.head.i == token.i
                dependencies.append({
                    **common,
                    "sentence_id": sentence_id,
                    "token_id": row["token_id"],
                    "surface": row["surface"],
                    "lemma": lemma,
                    "upos": row["upos"],
                    "dependency": str(token.dep_ or ""),
                    "head_token_id": 0 if head_is_root else token_ids.get(token.head.i, 0),
                    "head_surface": "" if head_is_root else str(token.head.text),
                    "named_entity": str(token.ent_type_ or ""),
                })
    return morphemes, dependencies, sentence_count


def _analyze_with_sudachi(
    segments: list[dict[str, Any]],
    dictionary: Any,
    split_mode: str,
    stop_words: set[str],
) -> tuple[list[dict[str, Any]], int]:
    morphemes: list[dict[str, Any]] = []
    sentence_count = 0
    with _ENGINE_LOCK:
        tokenizer = dictionary.create(mode=split_mode)
        for segment in segments:
            text = str(segment.get("text") or "")
            if not text:
                continue
            common = _row_common(segment)
            sentence_id = 1
            sentence_token_id = 0
            sentence_count += 1
            cursor = 0
            for token_id, token in enumerate(tokenizer.tokenize(text), 1):
                surface = str(token.surface())
                pos = tuple(str(value) for value in token.part_of_speech())
                upos = _upos_from_sudachi(pos)
                normalized = str(token.normalized_form() or token.dictionary_form() or surface)
                lemma = str(token.dictionary_form() or normalized or surface)
                try:
                    begin = int(token.begin())
                    end = int(token.end())
                except (AttributeError, TypeError, ValueError):
                    begin = text.find(surface, cursor)
                    begin = cursor if begin < 0 else begin
                    end = begin + len(surface)
                cursor = end
                sentence_token_id += 1
                is_stop = normalized.casefold() in stop_words
                morphemes.append({
                    **common,
                    "sentence_id": sentence_id,
                    "token_id": token_id,
                    "sentence_token_id": sentence_token_id,
                    "surface": surface,
                    "lemma": lemma,
                    "normalized": normalized,
                    "reading": str(token.reading_form() or ""),
                    "upos": upos,
                    "pos_detail": "-".join(value for value in pos if value and value != "*"),
                    "inflection": "|".join(
                        value for value in pos[4:] if value and value != "*"
                    ),
                    "begin": begin,
                    "end": end,
                    "is_content": pos[:1] in (("名詞",), ("動詞",), ("形容詞",), ("副詞",))
                    and not is_stop,
                    "is_stop": is_stop,
                })
                if SENTENCE_END_PATTERN.search(surface):
                    sentence_id += 1
                    sentence_token_id = 0
                    sentence_count += 1
            if morphemes and sentence_token_id == 0:
                sentence_count = max(0, sentence_count - 1)
    return morphemes, sentence_count


def _analyze_with_fallback(
    segments: list[dict[str, Any]],
    stop_words: set[str],
) -> tuple[list[dict[str, Any]], int]:
    morphemes: list[dict[str, Any]] = []
    sentence_count = 0
    for segment in segments:
        text = str(segment.get("text") or "")
        if not text:
            continue
        common = _row_common(segment)
        sentence_id = 1
        sentence_token_id = 0
        sentence_count += 1
        for token_id, match in enumerate(FALLBACK_TOKEN_PATTERN.finditer(text), 1):
            surface = match.group(0)
            sentence_token_id += 1
            if re.fullmatch(r"[^\w一-龯ぁ-ゖァ-ヺ々〆ヵヶー]+", surface):
                upos = "PUNCT"
            elif re.fullmatch(r"[0-9]+(?:[.,][0-9]+)*", surface):
                upos = "NUM"
            else:
                upos = "X"
            normalized = surface.casefold() if surface.isascii() else surface
            is_stop = normalized in stop_words
            morphemes.append({
                **common,
                "sentence_id": sentence_id,
                "token_id": token_id,
                "sentence_token_id": sentence_token_id,
                "surface": surface,
                "lemma": normalized,
                "normalized": normalized,
                "reading": "",
                "upos": upos,
                "pos_detail": "簡易文字種分割",
                "inflection": "",
                "begin": match.start(),
                "end": match.end(),
                "is_content": upos == "X" and not is_stop,
                "is_stop": is_stop,
            })
            if SENTENCE_END_PATTERN.search(surface):
                sentence_id += 1
                sentence_token_id = 0
                sentence_count += 1
        if morphemes and sentence_token_id == 0:
            sentence_count = max(0, sentence_count - 1)
    return morphemes, sentence_count


def _linguistic_analysis(
    segments: list[dict[str, Any]],
    config: dict[str, Any],
) -> dict[str, Any]:
    split_mode = str(config.get("morph_split_mode") or "C").upper()
    if split_mode not in {"A", "B", "C"}:
        split_mode = "C"
    stop_words = _clean_stop_words(config)
    ginza, ginza_error = _load_ginza()
    dependencies: list[dict[str, Any]] = []
    if ginza is not None:
        try:
            morphemes, dependencies, sentence_count = _analyze_with_ginza(
                segments, ginza, stop_words
            )
            engine = {
                "morphology": "GiNZA / SudachiPy",
                "syntax": "GiNZA Universal Dependencies",
                "version": _package_version("ginza"),
                "status": "ready",
                "split_mode": "GiNZAモデル設定",
                "message": "形態素情報と係り受けをGiNZAで解析しました。",
            }
        except Exception as error:
            ginza_error = _safe_error(error)
            ginza = None
    if ginza is None:
        dictionary, sudachi_error = _load_sudachi()
        if dictionary is not None:
            try:
                morphemes, sentence_count = _analyze_with_sudachi(
                    segments, dictionary, split_mode, stop_words
                )
                engine = {
                    "morphology": "SudachiPy",
                    "syntax": "利用不可（GiNZAモデル未読込）",
                    "version": _package_version("SudachiPy"),
                    "status": "partial",
                    "split_mode": split_mode,
                    "message": (
                        "形態素解析はSudachiPyで実行しました。係り受け解析には"
                        "ginza と ja-ginza の正常な導入が必要です。"
                    ),
                    "diagnostic": ginza_error,
                }
            except Exception as error:
                sudachi_error = _safe_error(error)
                dictionary = None
        if dictionary is None:
            morphemes, sentence_count = _analyze_with_fallback(segments, stop_words)
            engine = {
                "morphology": "簡易文字種分割（代替表示）",
                "syntax": "利用不可",
                "version": "",
                "status": "fallback",
                "split_mode": "",
                "message": (
                    "解析器を読み込めなかったため簡易分割を表示しています。"
                    "これは形態素解析・構文解析の研究データとして使用できません。"
                ),
                "diagnostic": "; ".join(
                    value for value in (ginza_error, sudachi_error) if value
                )[:500],
            }

    included_rows = [row for row in morphemes if not row["excluded"]]
    pos_counter = Counter(
        (str(row["upos"]), str(row["pos_detail"])) for row in included_rows
    )
    pos_total = sum(pos_counter.values())
    pos_frequency = [
        {
            "upos": upos,
            "pos_detail": detail,
            "count": count,
            "percent": round(100 * count / pos_total, 4) if pos_total else 0.0,
        }
        for (upos, detail), count in sorted(
            pos_counter.items(), key=lambda item: (-item[1], item[0])
        )
    ]

    documents: dict[str, set[str]] = defaultdict(set)
    speaker_terms: dict[str, set[str]] = defaultdict(set)
    term_counts: Counter[str] = Counter()
    term_pos: dict[str, Counter[str]] = defaultdict(Counter)
    term_lemma: dict[str, Counter[str]] = defaultdict(Counter)
    speaker_term_counts: dict[str, Counter[str]] = defaultdict(Counter)
    speaker_term_documents: dict[str, dict[str, set[str]]] = defaultdict(
        lambda: defaultdict(set)
    )
    speaker_document_ids: dict[str, set[str]] = defaultdict(set)
    speaker_term_meta: dict[str, dict[str, str]] = {}
    for row in included_rows:
        if not row["is_content"] or row["is_stop"]:
            continue
        term = str(row["normalized"] or row["lemma"] or row["surface"]).strip()
        if not term or re.fullmatch(r"[\W_]+", term, re.UNICODE):
            continue
        term = term.casefold() if term.isascii() else term
        term_counts[term] += 1
        term_pos[term][str(row["upos"])] += 1
        term_lemma[term][str(row["lemma"] or term)] += 1
        documents[str(row["segment_id"])].add(term)
        speaker = str(row["speaker"])
        segment_id = str(row["segment_id"])
        speaker_terms[speaker].add(term)
        speaker_term_counts[speaker][term] += 1
        speaker_term_documents[speaker][term].add(segment_id)
        speaker_document_ids[speaker].add(segment_id)
        speaker_term_meta.setdefault(speaker, {
            "speaker": speaker,
            "speaker_name": str(row.get("speaker_name") or speaker),
            "role": str(row.get("role") or "participant"),
        })
    document_frequency: Counter[str] = Counter()
    for terms in documents.values():
        document_frequency.update(terms)
    speaker_frequency: Counter[str] = Counter()
    for terms in speaker_terms.values():
        speaker_frequency.update(terms)
    document_count = len({
        str(segment.get("id") or "")
        for segment in segments if not segment.get("excluded") and str(segment.get("text") or "")
    })
    total_terms = sum(term_counts.values())
    term_frequency = []
    for rank, (term, count) in enumerate(
        sorted(term_counts.items(), key=lambda item: (-item[1], item[0])), 1
    ):
        df = int(document_frequency[term])
        inverse_document_frequency = math.log(
            (1 + max(1, document_count)) / (1 + df)
        ) + 1
        term_frequency.append({
            "rank": rank,
            "term": term,
            "lemma": term_lemma[term].most_common(1)[0][0],
            "upos": term_pos[term].most_common(1)[0][0],
            "term_frequency": int(count),
            "document_frequency": df,
            "speaker_frequency": int(speaker_frequency[term]),
            "term_percent": round(100 * count / total_terms, 5) if total_terms else 0.0,
            "document_percent": round(100 * df / document_count, 5)
            if document_count else 0.0,
            "tf_idf": round(count * inverse_document_frequency, 6),
        })

    speaker_term_frequency = []
    for speaker, counts in sorted(
        speaker_term_counts.items(),
        key=lambda item: (
            str(speaker_term_meta.get(item[0], {}).get("speaker_name") or item[0]),
            item[0],
        ),
    ):
        speaker_total = sum(counts.values())
        speaker_document_count = len(speaker_document_ids[speaker])
        terms = []
        for rank, (term, count) in enumerate(
            sorted(counts.items(), key=lambda item: (-item[1], item[0]))[:50], 1
        ):
            document_frequency_value = len(speaker_term_documents[speaker][term])
            terms.append({
                "rank": rank,
                "term": term,
                "lemma": term_lemma[term].most_common(1)[0][0],
                "upos": term_pos[term].most_common(1)[0][0],
                "term_frequency": int(count),
                "document_frequency": document_frequency_value,
                "term_percent": round(100 * count / speaker_total, 5)
                if speaker_total else 0.0,
                "document_percent": round(
                    100 * document_frequency_value / speaker_document_count, 5
                ) if speaker_document_count else 0.0,
            })
        speaker_term_frequency.append({
            **speaker_term_meta.get(speaker, {
                "speaker": speaker,
                "speaker_name": speaker,
                "role": "participant",
            }),
            "document_count": speaker_document_count,
            "content_token_count": speaker_total,
            "unique_term_count": len(counts),
            "terms": terms,
        })

    top_terms = int(config.get("cooccurrence_top_terms") or 60)
    top_terms = min(200, max(10, top_terms))
    minimum_count = int(config.get("cooccurrence_min_count") or 2)
    minimum_count = min(1000, max(1, minimum_count))
    eligible = {row["term"] for row in term_frequency[:top_terms]}
    pair_counts: Counter[tuple[str, str]] = Counter()
    for terms in documents.values():
        selected = sorted(terms & eligible)
        pair_counts.update(combinations(selected, 2))
    cooccurrence = []
    for pair, count in pair_counts.items():
        if count < minimum_count:
            continue
        term_a, term_b = pair
        df_a = document_frequency[term_a]
        df_b = document_frequency[term_b]
        union = df_a + df_b - count
        cooccurrence.append({
            "term_a": term_a,
            "term_b": term_b,
            "cooccurrence_count": int(count),
            "document_frequency_a": int(df_a),
            "document_frequency_b": int(df_b),
            "jaccard": round(count / union, 6) if union else 0.0,
            "dice": round(2 * count / (df_a + df_b), 6) if df_a + df_b else 0.0,
        })
    cooccurrence.sort(
        key=lambda row: (-row["jaccard"], -row["cooccurrence_count"], row["term_a"], row["term_b"])
    )
    cooccurrence = [
        {"rank": rank, **row} for rank, row in enumerate(cooccurrence[:500], 1)
    ]

    return {
        "engine": engine,
        "coverage": {
            "segment_count": len(segments),
            "included_segment_count": sum(
                1 for segment in segments if not segment.get("excluded")
            ),
            "sentence_count": sentence_count,
            "token_count": len(morphemes),
            "included_token_count": len(included_rows),
            "content_token_count": sum(1 for row in included_rows if row["is_content"]),
            "dependency_count": len(dependencies),
            "document_unit": "発話",
            "cooccurrence_min_count": minimum_count,
            "cooccurrence_top_terms": top_terms,
        },
        "morphemes": morphemes,
        "dependencies": dependencies,
        "pos_frequency": pos_frequency,
        "term_frequency": term_frequency,
        "speaker_term_frequency": speaker_term_frequency,
        "cooccurrence": cooccurrence,
    }


NUMERIC_VARIABLES: dict[str, str] = {
    "duration_seconds": "発話時間（秒）",
    "characters": "文字数",
    "token_count": "形態素数",
    "content_token_count": "内容語数",
    "lexical_diversity": "語彙多様性（異なり内容語/内容語）",
    "characters_per_minute": "発話速度（文字/分）",
}


def _segment_dataset(
    analysis: dict[str, Any],
    morphemes: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    tokens_by_segment: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in morphemes:
        tokens_by_segment[str(row["segment_id"])].append(row)
    code_labels = {
        str(item.get("id")): str(item.get("label") or item.get("id"))
        for item in analysis.get("manual", {}).get("codebook", [])
    }
    rows = []
    for segment in analysis.get("segments", []):
        segment_id = str(segment.get("id") or "")
        token_rows = tokens_by_segment.get(segment_id, [])
        content_terms = [
            str(row["normalized"] or row["lemma"] or row["surface"])
            for row in token_rows if row["is_content"] and not row["is_stop"]
        ]
        normalized_terms = {
            (term.casefold() if term.isascii() else term)
            for row in token_rows
            for term in (
                str(row.get("normalized") or "").strip(),
                str(row.get("lemma") or "").strip(),
                str(row.get("surface") or "").strip(),
            )
            if term
        }
        duration = max(0.0, float(segment.get("duration") or 0))
        characters = int(segment.get("characters") or 0)
        annotation = segment.get("annotation") if isinstance(segment.get("annotation"), dict) else {}
        code_ids = [str(value) for value in annotation.get("codes", [])]
        rows.append({
            "segment_id": segment_id,
            "start": segment.get("start", 0),
            "end": segment.get("end", 0),
            "duration_seconds": round(duration, 3),
            "speaker": segment.get("speaker", "UNKNOWN"),
            "speaker_name": segment.get("speaker_name", ""),
            "role": segment.get("role", "participant"),
            "text": segment.get("text", ""),
            "characters": characters,
            "token_count": len(token_rows),
            "content_token_count": len(content_terms),
            "unique_content_terms": len(set(content_terms)),
            "lexical_diversity": round(len(set(content_terms)) / len(content_terms), 6)
            if content_terms else None,
            "characters_per_minute": round(60 * characters / duration, 4)
            if duration > 0 else None,
            "question_candidate": bool(segment.get("question_candidate")),
            "code_ids": code_ids,
            "code_labels": [code_labels.get(value, value) for value in code_ids],
            "interaction_tags": annotation.get("interaction_tags", []),
            "important": bool(annotation.get("important")),
            "excluded": bool(segment.get("excluded")),
            "_normalized_terms": normalized_terms,
        })
    return rows


def _descriptive_row(
    values: list[float],
    *,
    scope: str,
    group_variable: str,
    group: str,
    variable: str,
    missing: int,
) -> dict[str, Any]:
    return {
        "scope": scope,
        "group_variable": group_variable,
        "group": group,
        "variable": variable,
        "label": NUMERIC_VARIABLES[variable],
        "n": len(values),
        "missing": missing,
        "mean": _finite(mean(values)) if values else None,
        "standard_deviation": _finite(stdev(values)) if len(values) > 1 else None,
        "median": _finite(median(values)) if values else None,
        "minimum": _finite(min(values)) if values else None,
        "q1": _finite(_percentile(values, 0.25)),
        "q3": _finite(_percentile(values, 0.75)),
        "maximum": _finite(max(values)) if values else None,
    }


def _frequency_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result = []
    variables: list[tuple[str, str, Callable[[dict[str, Any]], str]]] = [
        ("speaker", "話者", lambda row: str(row["speaker_name"] or row["speaker"])),
        ("role", "役割", lambda row: str(row["role"] or "未設定")),
        (
            "question_candidate",
            "質問候補",
            lambda row: "質問候補" if row["question_candidate"] else "その他",
        ),
    ]
    for variable, label, accessor in variables:
        counts = Counter(accessor(row) for row in rows)
        total = sum(counts.values())
        for value, count in sorted(counts.items(), key=lambda item: (-item[1], item[0])):
            result.append({
                "variable": variable,
                "label": label,
                "value": value,
                "count": count,
                "percent": round(100 * count / total, 5) if total else 0.0,
            })
    return result


def _crosstab_rows(
    rows: list[dict[str, Any]],
    *,
    table_id: str,
    table_label: str,
    group_variable: str,
    group_accessor: Callable[[dict[str, Any]], str],
    column_variable: str,
    column_accessor: Callable[[dict[str, Any]], str],
) -> tuple[list[dict[str, Any]], list[str], list[str], list[list[int]]]:
    row_values = sorted({group_accessor(row) for row in rows})
    column_values = sorted({column_accessor(row) for row in rows})
    counts = Counter((group_accessor(row), column_accessor(row)) for row in rows)
    row_totals = Counter()
    column_totals = Counter()
    for (row_value, column_value), count in counts.items():
        row_totals[row_value] += count
        column_totals[column_value] += count
    total = sum(counts.values())
    result = []
    matrix = []
    for row_value in row_values:
        matrix_row = []
        for column_value in column_values:
            count = int(counts[(row_value, column_value)])
            matrix_row.append(count)
            result.append({
                "table_id": table_id,
                "table_label": table_label,
                "row_variable": group_variable,
                "row_value": row_value,
                "column_variable": column_variable,
                "column_value": column_value,
                "count": count,
                "row_percent": round(100 * count / row_totals[row_value], 5)
                if row_totals[row_value] else 0.0,
                "column_percent": round(100 * count / column_totals[column_value], 5)
                if column_totals[column_value] else 0.0,
                "total_percent": round(100 * count / total, 5) if total else 0.0,
            })
        matrix.append(matrix_row)
    return result, row_values, column_values, matrix


def _test_result_base(
    family: str,
    test: str,
    outcome: str,
    outcome_label: str,
    group_variable: str,
    n: int,
    groups: int,
) -> dict[str, Any]:
    return {
        "family": family,
        "test": test,
        "outcome": outcome,
        "outcome_label": outcome_label,
        "group_variable": group_variable,
        "n": n,
        "groups": groups,
        "statistic": None,
        "df1": None,
        "df2": None,
        "p_value": None,
        "effect_name": "",
        "effect_size": None,
        "significant_0_05": False,
        "status": "not_computable",
        "assumption_note": "",
    }


def _statistics_analysis(
    analysis: dict[str, Any],
    segment_rows: list[dict[str, Any]],
) -> dict[str, Any]:
    included = [row for row in segment_rows if not row["excluded"]]
    config = analysis.get("config") if isinstance(analysis.get("config"), dict) else {}
    group_variable = str(config.get("statistics_group_by") or "speaker")
    if group_variable not in {"speaker", "role"}:
        group_variable = "speaker"

    def group_value(row: dict[str, Any]) -> str:
        if group_variable == "role":
            return str(row["role"] or "未設定")
        return str(row["speaker_name"] or row["speaker"] or "話者未判定")

    descriptives = []
    for variable in NUMERIC_VARIABLES:
        values = [
            float(row[variable]) for row in included
            if row.get(variable) is not None and math.isfinite(float(row[variable]))
        ]
        descriptives.append(_descriptive_row(
            values,
            scope="overall",
            group_variable=group_variable,
            group="全体",
            variable=variable,
            missing=len(included) - len(values),
        ))
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in included:
        groups[group_value(row)].append(row)
    for group, group_rows in sorted(groups.items()):
        for variable in NUMERIC_VARIABLES:
            values = [
                float(row[variable]) for row in group_rows
                if row.get(variable) is not None and math.isfinite(float(row[variable]))
            ]
            descriptives.append(_descriptive_row(
                values,
                scope="group",
                group_variable=group_variable,
                group=group,
                variable=variable,
                missing=len(group_rows) - len(values),
            ))

    frequencies = _frequency_rows(included)
    crosstabs = []
    contingency_inputs: list[tuple[str, str, list[str], list[str], list[list[int]]]] = []
    question_rows, row_values, column_values, matrix = _crosstab_rows(
        included,
        table_id=f"{group_variable}_x_question",
        table_label=f"{group_variable} × 質問候補",
        group_variable=group_variable,
        group_accessor=group_value,
        column_variable="question_candidate",
        column_accessor=lambda row: "質問候補" if row["question_candidate"] else "その他",
    )
    crosstabs.extend(question_rows)
    contingency_inputs.append((
        f"{group_variable}_x_question",
        "質問候補",
        row_values,
        column_values,
        matrix,
    ))

    codebook = analysis.get("manual", {}).get("codebook", [])
    for code in codebook[:50]:
        code_id = str(code.get("id") or "")
        code_label = str(code.get("label") or code_id)
        table_id = f"{group_variable}_x_code_{code_id}"
        table_rows, row_values, column_values, matrix = _crosstab_rows(
            included,
            table_id=table_id,
            table_label=f"{group_variable} × コード「{code_label}」",
            group_variable=group_variable,
            group_accessor=group_value,
            column_variable=f"code:{code_id}",
            column_accessor=lambda row, target=code_id: (
                "あり" if target in row.get("code_ids", []) else "なし"
            ),
        )
        crosstabs.extend(table_rows)
        contingency_inputs.append((
            table_id, f"コード「{code_label}」", row_values, column_values, matrix
        ))

    selected_terms = []
    for raw_term in config.get("crosstab_terms") or []:
        term = str(raw_term).strip()
        if term and term not in selected_terms:
            selected_terms.append(term)
        if len(selected_terms) >= 30:
            break

    def selected_term_present(row: dict[str, Any], term: str) -> bool:
        normalized = term.casefold() if term.isascii() else term
        if normalized in row.get("_normalized_terms", set()):
            return True
        text = str(row.get("text") or "")
        searchable = text.casefold() if term.isascii() else text
        return normalized in searchable

    for index, term in enumerate(selected_terms, 1):
        table_id = f"{group_variable}_x_selected_term_{index}"
        table_rows, row_values, column_values, matrix = _crosstab_rows(
            included,
            table_id=table_id,
            table_label=f"{group_variable} × 単語「{term}」",
            group_variable=group_variable,
            group_accessor=group_value,
            column_variable=f"selected_term:{term}",
            column_accessor=lambda row, target=term: (
                "あり" if selected_term_present(row, target) else "なし"
            ),
        )
        crosstabs.extend(table_rows)
        contingency_inputs.append((
            table_id, f"単語「{term}」", row_values, column_values, matrix
        ))

    tests: list[dict[str, Any]] = []
    correlations: list[dict[str, Any]] = []
    scipy_error = ""
    try:
        from scipy import stats
    except Exception as error:
        stats = None
        scipy_error = _safe_error(error)

    for variable, label in NUMERIC_VARIABLES.items():
        samples = []
        for group_rows in groups.values():
            values = [
                float(row[variable]) for row in group_rows
                if row.get(variable) is not None and math.isfinite(float(row[variable]))
            ]
            if values:
                samples.append(values)
        n = sum(len(sample) for sample in samples)
        anova = _test_result_base(
            "平均値比較", "一元配置分散分析（ANOVA）", variable, label,
            group_variable, n, len(samples),
        )
        anova["effect_name"] = "eta_squared"
        anova["assumption_note"] = (
            "正規性、等分散性、観測の独立性を要確認。発話は同一会話・話者内で"
            "独立でない可能性が高いため探索的に扱ってください。"
        )
        if (
            stats is not None
            and len(samples) >= 2
            and all(len(sample) >= 2 for sample in samples)
            and len({value for sample in samples for value in sample}) > 1
        ):
            try:
                statistic, p_value = stats.f_oneway(*samples)
                grand = mean([value for sample in samples for value in sample])
                ss_between = sum(
                    len(sample) * (mean(sample) - grand) ** 2 for sample in samples
                )
                ss_total = sum(
                    (value - grand) ** 2 for sample in samples for value in sample
                )
                anova.update({
                    "statistic": _finite(statistic),
                    "df1": len(samples) - 1,
                    "df2": n - len(samples),
                    "p_value": _finite(p_value),
                    "effect_size": _finite(ss_between / ss_total) if ss_total else 0.0,
                    "significant_0_05": bool(math.isfinite(p_value) and p_value < 0.05),
                    "status": "computed",
                })
            except Exception as error:
                anova["status"] = "error"
                anova["assumption_note"] += f" 計算エラー: {_safe_error(error)}"
        elif stats is None:
            anova["status"] = "unavailable"
            anova["assumption_note"] += f" SciPyを読み込めません: {scipy_error}"
        tests.append(anova)

        kruskal = _test_result_base(
            "分布比較", "Kruskal–Wallis検定", variable, label,
            group_variable, n, len(samples),
        )
        kruskal["effect_name"] = "epsilon_squared"
        kruskal["assumption_note"] = (
            "順位に基づく比較ですが、観測の独立性は必要です。発話の反復・"
            "話者内相関があるため探索的に扱ってください。"
        )
        if (
            stats is not None
            and len(samples) >= 2
            and all(len(sample) >= 2 for sample in samples)
            and len({value for sample in samples for value in sample}) > 1
        ):
            try:
                statistic, p_value = stats.kruskal(*samples)
                denominator = n - len(samples)
                epsilon_squared = (
                    max(0.0, (float(statistic) - len(samples) + 1) / denominator)
                    if denominator > 0 else 0.0
                )
                kruskal.update({
                    "statistic": _finite(statistic),
                    "df1": len(samples) - 1,
                    "p_value": _finite(p_value),
                    "effect_size": _finite(epsilon_squared),
                    "significant_0_05": bool(math.isfinite(p_value) and p_value < 0.05),
                    "status": "computed",
                })
            except Exception as error:
                kruskal["status"] = "error"
                kruskal["assumption_note"] += f" 計算エラー: {_safe_error(error)}"
        elif stats is None:
            kruskal["status"] = "unavailable"
            kruskal["assumption_note"] += f" SciPyを読み込めません: {scipy_error}"
        tests.append(kruskal)

    for table_id, label, row_values, column_values, matrix in contingency_inputs:
        n = sum(sum(row) for row in matrix)
        chi_square = _test_result_base(
            "クロス集計", "Pearsonのカイ二乗検定", table_id, label,
            group_variable, n, len(row_values),
        )
        chi_square["effect_name"] = "cramers_v"
        chi_square["assumption_note"] = (
            "期待度数5未満のセルが多い場合は漸近p値が不正確です。"
            "発話は独立観測でない可能性があるため探索的に扱ってください。"
        )
        nonzero_rows = [row for row in matrix if sum(row) > 0]
        nonzero_columns = [
            index for index in range(len(column_values))
            if sum(row[index] for row in nonzero_rows) > 0
        ]
        cleaned = [
            [row[index] for index in nonzero_columns] for row in nonzero_rows
        ]
        if stats is not None and len(cleaned) >= 2 and len(nonzero_columns) >= 2:
            try:
                statistic, p_value, dof, expected = stats.chi2_contingency(
                    cleaned, correction=False
                )
                minimum_dimension = min(len(cleaned) - 1, len(nonzero_columns) - 1)
                cramers_v = math.sqrt(statistic / (n * minimum_dimension)) \
                    if n > 0 and minimum_dimension > 0 else 0.0
                low_expected = sum(
                    1 for row in expected for value in row if float(value) < 5
                )
                expected_count = sum(len(row) for row in expected)
                chi_square.update({
                    "statistic": _finite(statistic),
                    "df1": int(dof),
                    "p_value": _finite(p_value),
                    "effect_size": _finite(cramers_v),
                    "significant_0_05": bool(math.isfinite(p_value) and p_value < 0.05),
                    "status": "computed",
                    "assumption_note": (
                        chi_square["assumption_note"]
                        + f" 期待度数5未満: {low_expected}/{expected_count}セル。"
                    ),
                })
            except Exception as error:
                chi_square["status"] = "error"
                chi_square["assumption_note"] += f" 計算エラー: {_safe_error(error)}"
        elif stats is None:
            chi_square["status"] = "unavailable"
            chi_square["assumption_note"] += f" SciPyを読み込めません: {scipy_error}"
        tests.append(chi_square)

    for variable_a, variable_b in combinations(NUMERIC_VARIABLES, 2):
        pairs = [
            (float(row[variable_a]), float(row[variable_b]))
            for row in included
            if row.get(variable_a) is not None and row.get(variable_b) is not None
            and math.isfinite(float(row[variable_a]))
            and math.isfinite(float(row[variable_b]))
        ]
        for method, function in (
            ("Pearson", stats.pearsonr if stats is not None else None),
            ("Spearman", stats.spearmanr if stats is not None else None),
        ):
            result = {
                "method": method,
                "variable_a": variable_a,
                "label_a": NUMERIC_VARIABLES[variable_a],
                "variable_b": variable_b,
                "label_b": NUMERIC_VARIABLES[variable_b],
                "n": len(pairs),
                "coefficient": None,
                "p_value": None,
                "significant_0_05": False,
                "status": "not_computable",
            }
            if function is None:
                result["status"] = "unavailable"
            elif (
                len(pairs) >= 3
                and len({pair[0] for pair in pairs}) > 1
                and len({pair[1] for pair in pairs}) > 1
            ):
                try:
                    coefficient, p_value = function(
                        [pair[0] for pair in pairs],
                        [pair[1] for pair in pairs],
                    )
                    result.update({
                        "coefficient": _finite(coefficient),
                        "p_value": _finite(p_value),
                        "significant_0_05": bool(
                            math.isfinite(float(p_value)) and float(p_value) < 0.05
                        ),
                        "status": "computed",
                    })
                except Exception:
                    result["status"] = "error"
            correlations.append(result)

    return {
        "analysis_unit": "発話",
        "group_variable": group_variable,
        "group_count": len(groups),
        "selected_terms": selected_terms,
        "descriptives": descriptives,
        "frequencies": frequencies,
        "crosstabs": crosstabs,
        "tests": tests,
        "correlations": correlations,
        "engine": {
            "name": "SciPy" if stats is not None else "利用不可",
            "version": _package_version("scipy"),
            "status": "ready" if stats is not None else "unavailable",
            "diagnostic": scipy_error,
        },
    }


def _method_rows(linguistics: dict[str, Any], statistics: dict[str, Any]) -> list[dict[str, Any]]:
    engine = linguistics["engine"]
    source_map = {row["id"]: row for row in SOURCE_REFERENCES}
    return [
        {
            "category": "形態素解析",
            "method": "表層形・辞書形・正規形・読み・品詞",
            "engine": engine["morphology"],
            "engine_version": engine.get("version", ""),
            "status": engine["status"],
            "analysis_unit": "形態素",
            "description": "GiNZA/SudachiPyの解析結果をトークン単位で保存します。",
            "source_title": source_map["sudachipy"]["title"],
            "source_url": source_map["sudachipy"]["url"],
        },
        {
            "category": "構文解析",
            "method": "Universal Dependencies係り受け",
            "engine": engine["syntax"],
            "engine_version": engine.get("version", ""),
            "status": "ready" if linguistics["dependencies"] else "unavailable",
            "analysis_unit": "形態素間の係り受け",
            "description": "各形態素の係り先と依存関係ラベルを保存します。",
            "source_title": source_map["ginza"]["title"],
            "source_url": source_map["ginza"]["url"],
        },
        {
            "category": "計量テキスト分析",
            "method": "TF・DF・Jaccard共起",
            "engine": "グルモジ内蔵集計",
            "engine_version": RESEARCH_ALGORITHM_VERSION,
            "status": "ready",
            "analysis_unit": "発話",
            "description": (
                "各発話を1文書として内容語のTF/DFを数え、同一発話内の"
                "共起をJaccard係数で集計します。KH Coder互換出力ではありません。"
            ),
            "source_title": source_map["khcoder"]["title"],
            "source_url": source_map["khcoder"]["url"],
        },
        {
            "category": "統計分析",
            "method": "記述統計・クロス集計・検定・相関",
            "engine": statistics["engine"]["name"],
            "engine_version": statistics["engine"]["version"],
            "status": statistics["engine"]["status"],
            "analysis_unit": "発話",
            "description": (
                "ANOVA、Kruskal–Wallis、Pearsonカイ二乗、Pearson/Spearman"
                "相関と効果量を探索的に計算します。"
            ),
            "source_title": source_map["scipy"]["title"],
            "source_url": source_map["scipy"]["url"],
        },
    ]


def _cache_key(analysis: dict[str, Any]) -> tuple[Any, ...]:
    item = analysis.get("item") if isinstance(analysis.get("item"), dict) else {}
    config = analysis.get("config") if isinstance(analysis.get("config"), dict) else {}
    return (
        item.get("id"),
        item.get("revision_count"),
        item.get("analysis_revision"),
        item.get("updated_at"),
        item.get("analysis_updated_at"),
        str(config.get("morph_split_mode") or "C"),
        int(config.get("cooccurrence_min_count") or 2),
        int(config.get("cooccurrence_top_terms") or 60),
        str(config.get("statistics_group_by") or "speaker"),
        tuple(str(value) for value in config.get("crosstab_terms", [])),
        tuple(sorted(str(value) for value in config.get("stop_words", []))),
    )


def build_research_analysis(analysis: dict[str, Any]) -> dict[str, Any]:
    key = _cache_key(analysis)
    with _ENGINE_LOCK:
        cached = _RESEARCH_CACHE.get(key)
        if cached is not None:
            _RESEARCH_CACHE.move_to_end(key)
            return cached
    linguistics = _linguistic_analysis(
        list(analysis.get("segments") or []),
        dict(analysis.get("config") or {}),
    )
    segment_rows = _segment_dataset(analysis, linguistics["morphemes"])
    statistics = _statistics_analysis(analysis, segment_rows)
    for row in segment_rows:
        row.pop("_normalized_terms", None)
    research = {
        "schema_version": 1,
        "algorithm_version": RESEARCH_ALGORITHM_VERSION,
        "analysis_unit": "発話",
        "methods": _method_rows(linguistics, statistics),
        "sources": SOURCE_REFERENCES,
        "limitations": [
            "自動解析には誤りが含まれます。重要語、テーマ、係り受けは原文と音声を確認してください。",
            "共起は同一発話内に現れた語の関連候補であり、因果関係や意味的関係を示しません。",
            "本実装のTF・DF・Jaccard共起はKH Coderの考え方を参照した独自集計で、KH Coderの結果との数値互換性はありません。",
            "同一会話内の発話は独立観測とは限りません。推測統計とp値は探索的に扱い、研究デザインに応じて話者・会話単位のモデルを別途検討してください。",
            "複数の検定に対する補正は行っていません。効果量、標本数、前提条件、研究上の意味を併記してください。",
        ],
        "linguistics": linguistics,
        "statistics": statistics,
        "segments": segment_rows,
    }
    with _ENGINE_LOCK:
        _RESEARCH_CACHE[key] = research
        _RESEARCH_CACHE.move_to_end(key)
        while len(_RESEARCH_CACHE) > _RESEARCH_CACHE_SIZE:
            _RESEARCH_CACHE.popitem(last=False)
    return research


def enrich_research_analysis(
    analysis: dict[str, Any],
    *,
    include_rows: bool = False,
) -> dict[str, Any]:
    research = build_research_analysis(analysis)
    linguistics = research["linguistics"]
    public_linguistics = {
        "engine": linguistics["engine"],
        "coverage": linguistics["coverage"],
        "pos_frequency": linguistics["pos_frequency"],
        "term_frequency": linguistics["term_frequency"][:200],
        "speaker_term_frequency": linguistics["speaker_term_frequency"],
        "cooccurrence": linguistics["cooccurrence"][:200],
        "morpheme_preview": linguistics["morphemes"][:100],
        "dependency_preview": linguistics["dependencies"][:100],
    }
    public_research = {
        **research,
        "linguistics": public_linguistics,
        "segments": research["segments"] if include_rows else [],
    }
    if include_rows:
        public_linguistics["morphemes"] = linguistics["morphemes"]
        public_linguistics["dependencies"] = linguistics["dependencies"]
    analysis["research"] = public_research
    exports = analysis.setdefault("exports", {})
    item_id = analysis.get("item", {}).get("id", "")
    exports.update({
        "xlsx": f"/api/library/{item_id}/analysis/export.xlsx",
        "segments_all": f"/api/library/{item_id}/analysis/export.csv?dataset=segments_all",
        "morphemes": f"/api/library/{item_id}/analysis/export.csv?dataset=morphemes",
        "dependencies": f"/api/library/{item_id}/analysis/export.csv?dataset=dependencies",
        "pos_frequency": f"/api/library/{item_id}/analysis/export.csv?dataset=pos_frequency",
        "term_frequency": f"/api/library/{item_id}/analysis/export.csv?dataset=term_frequency",
        "cooccurrence": f"/api/library/{item_id}/analysis/export.csv?dataset=cooccurrence",
        "descriptives": f"/api/library/{item_id}/analysis/export.csv?dataset=descriptives",
        "frequencies": f"/api/library/{item_id}/analysis/export.csv?dataset=frequencies",
        "crosstabs": f"/api/library/{item_id}/analysis/export.csv?dataset=crosstabs",
        "statistical_tests": f"/api/library/{item_id}/analysis/export.csv?dataset=statistical_tests",
        "correlations": f"/api/library/{item_id}/analysis/export.csv?dataset=correlations",
        "analysis_methods": f"/api/library/{item_id}/analysis/export.csv?dataset=analysis_methods",
    })
    analysis.setdefault("classification", {}).setdefault("automatic", []).extend([
        "形態素解析", "係り受け解析", "語彙頻度・共起", "記述統計・探索的検定",
    ])
    for caution in research["limitations"]:
        if caution not in analysis.setdefault("cautions", []):
            analysis["cautions"].append(caution)
    return analysis


def research_csv_sources(analysis: dict[str, Any]) -> dict[str, list[dict[str, Any]]]:
    research = analysis.get("research") if isinstance(analysis.get("research"), dict) else {}
    linguistics = (
        research.get("linguistics")
        if isinstance(research.get("linguistics"), dict)
        else {}
    )
    statistics = (
        research.get("statistics")
        if isinstance(research.get("statistics"), dict)
        else {}
    )
    return {
        "segments_all": list(research.get("segments") or []),
        "morphemes": list(linguistics.get("morphemes") or []),
        "dependencies": list(linguistics.get("dependencies") or []),
        "pos_frequency": list(linguistics.get("pos_frequency") or []),
        "term_frequency": list(linguistics.get("term_frequency") or []),
        "cooccurrence": list(linguistics.get("cooccurrence") or []),
        "descriptives": list(statistics.get("descriptives") or []),
        "frequencies": list(statistics.get("frequencies") or []),
        "crosstabs": list(statistics.get("crosstabs") or []),
        "statistical_tests": list(statistics.get("tests") or []),
        "correlations": list(statistics.get("correlations") or []),
        "analysis_methods": list(research.get("methods") or []),
    }


def _excel_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, float) and not math.isfinite(value):
        return ""
    if isinstance(value, (dict, list, tuple, set)):
        value = json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    if not isinstance(value, str):
        return value
    value = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", value)
    if value.lstrip().startswith(EXCEL_FORMULA_PREFIXES):
        return "'" + value
    return value


EXCEL_SHEETS = [
    ("segments_all", "発話データ"),
    ("speakers", "話者別集計"),
    ("morphemes", "形態素"),
    ("dependencies", "構文・係り受け"),
    ("pos_frequency", "品詞頻度"),
    ("term_frequency", "語彙頻度"),
    ("cooccurrence", "共起"),
    ("descriptives", "記述統計"),
    ("frequencies", "度数分布"),
    ("crosstabs", "クロス集計"),
    ("statistical_tests", "統計検定"),
    ("correlations", "相関"),
    ("summary", "既存分析概要"),
    ("observations", "確認候補"),
    ("transitions", "話者遷移"),
    ("gaps", "無音候補"),
    ("overlaps", "発話重なり"),
    ("timeline", "時間推移"),
    ("emotions", "感情推定"),
    ("groups", "属性比較"),
    ("codes", "コード集計"),
    ("coded_segments", "コード済み発話"),
    ("interactions", "相互作用"),
    ("case_matrix", "話者コード表"),
    ("important_quotes", "重要引用"),
    ("context", "確認状況"),
    ("analysis_methods", "分析手法"),
]


def build_analysis_workbook(
    analysis: dict[str, Any],
    datasets: dict[str, tuple[list[str], list[dict[str, Any]]]],
) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception as error:
        raise RuntimeError(
            "Excel出力にはopenpyxlが必要です。run.batを再実行して依存関係を更新してください。"
        ) from error

    workbook = Workbook()
    workbook.properties.title = "グルモジ 研究分析データ"
    workbook.properties.creator = "グルモジ"
    workbook.properties.description = (
        "文字起こし、形態素・構文解析、計量テキスト分析、統計分析"
    )
    readme = workbook.active
    readme.title = "README"
    title_fill = PatternFill("solid", fgColor="1C6B50")
    header_fill = PatternFill("solid", fgColor="DCEADF")
    caution_fill = PatternFill("solid", fgColor="FFF2CC")
    readme["A1"] = "グルモジ 研究分析ワークブック"
    readme["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    readme["A1"].fill = title_fill
    readme.merge_cells("A1:D1")
    metadata_rows = [
        ("対象ID", analysis.get("item", {}).get("id", "")),
        ("元ファイル", analysis.get("item", {}).get("source_name", "")),
        ("生成日時", analysis.get("generated_at", "")),
        ("文字起こしrevision", analysis.get("item", {}).get("revision_count", "")),
        ("分析revision", analysis.get("item", {}).get("analysis_revision", "")),
        ("標準分析単位", "発話"),
        ("文字コード", "Excel Open XML (.xlsx)"),
    ]
    row_index = 3
    for label, value in metadata_rows:
        readme.cell(row_index, 1, label).font = Font(bold=True)
        readme.cell(row_index, 2, _excel_value(value))
        row_index += 1
    row_index += 1
    readme.cell(row_index, 1, "重要な注意").font = Font(bold=True)
    readme.cell(row_index, 1).fill = caution_fill
    row_index += 1
    limitations = analysis.get("research", {}).get("limitations", [])
    for limitation in limitations:
        readme.cell(row_index, 1, "・" + str(limitation))
        readme.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=4)
        readme.cell(row_index, 1).alignment = Alignment(wrap_text=True, vertical="top")
        row_index += 1
    row_index += 1
    readme.cell(row_index, 1, "分析手法と出典").font = Font(bold=True)
    readme.cell(row_index, 1).fill = header_fill
    row_index += 1
    method_headers = ["区分", "手法", "エンジン", "状態", "出典URL"]
    for column, value in enumerate(method_headers, 1):
        cell = readme.cell(row_index, column, value)
        cell.font = Font(bold=True)
        cell.fill = header_fill
    row_index += 1
    for method in analysis.get("research", {}).get("methods", []):
        values = [
            method.get("category"), method.get("method"), method.get("engine"),
            method.get("status"), method.get("source_url"),
        ]
        for column, value in enumerate(values, 1):
            readme.cell(row_index, column, _excel_value(value))
        row_index += 1
    for column, width in enumerate((24, 55, 30, 16, 55), 1):
        readme.column_dimensions[get_column_letter(column)].width = width
    readme.freeze_panes = "A3"

    for dataset, sheet_name in EXCEL_SHEETS:
        fields, rows = datasets.get(dataset, (RESEARCH_CSV_FIELDS.get(dataset, []), []))
        if len(rows) > 1_048_575:
            raise ValueError(
                f"{sheet_name}はExcelの行数上限を超えています。CSVを使用してください。"
            )
        worksheet = workbook.create_sheet(sheet_name[:31])
        for column, field in enumerate(fields, 1):
            cell = worksheet.cell(1, column, field)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = title_fill
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        for row_number, row in enumerate(rows, 2):
            for column, field in enumerate(fields, 1):
                cell = worksheet.cell(row_number, column, _excel_value(row.get(field)))
                cell.alignment = Alignment(vertical="top", wrap_text=field in {
                    "text", "description", "message", "assumption_note", "memo",
                })
        worksheet.freeze_panes = "A2"
        if fields:
            worksheet.auto_filter.ref = (
                f"A1:{get_column_letter(len(fields))}{max(1, len(rows) + 1)}"
            )
        for column, field in enumerate(fields, 1):
            sample_values = [str(field)]
            sample_values.extend(
                str(_excel_value(row.get(field, ""))) for row in rows[:200]
            )
            width = min(60, max(10, max((len(value) for value in sample_values), default=10) + 2))
            if field in {"text", "description", "message", "assumption_note", "memo"}:
                width = min(60, max(width, 32))
            worksheet.column_dimensions[get_column_letter(column)].width = width

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()
